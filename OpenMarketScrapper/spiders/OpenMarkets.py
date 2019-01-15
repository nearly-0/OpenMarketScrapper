import scrapy
from openpyxl import load_workbook
from openpyxl import Workbook
import random

class ProductInfo(scrapy.Item):
    # define the fields for your item here like:
    title = scrapy.Field()
    price = scrapy.Field()
    link = scrapy.Field()

    pass

    #################################### https pattern ##################################################################################################################################################################################################################
    # 'https://search.shopping.naver.com/search/all.nhn?origQuery=%EC%9D%B8%EC%B2%B4%EB%AA%A8%ED%98%95%20XC-105&pagingIndex=1&pagingSize=80&viewType=list&sort=rel&frm=NVSHATC&query=%EC%9D%B8%EC%B2%B4%EB%AA%A8%ED%98%95%20XC-105'
    # 'https://search.shopping.naver.com/search/all.nhn?origQuery=%EC%9D%B8%EC%B2%B4%EB%AA%A8%ED%98%95%20XC-110&pagingIndex=1&pagingSize=80&viewType=list&sort=rel&frm=NVSHATC&query=%EC%9D%B8%EC%B2%B4%EB%AA%A8%ED%98%95%20XC-110'
    # 'https://search.shopping.naver.com/search/all.nhn?origQuery=%EC%9D%B8%EC%B2%B4%EB%AA%A8%ED%98%95%20XC-111&pagingIndex=1&pagingSize=80&viewType=list&sort=rel&frm=NVSHATC&query=%EC%9D%B8%EC%B2%B4%EB%AA%A8%ED%98%95%20XC-111'
    # 'https://search.shopping.naver.com/search/all.nhn?origQuery=%EC%9D%B8%EC%B2%B4%EB%AA%A8%ED%98%95%20XC-401&pagingIndex=1&pagingSize=80&viewType=list&sort=rel&frm=NVSHATC&query=%EC%9D%B8%EC%B2%B4%EB%AA%A8%ED%98%95%20XC-401'

    # https://search.shopping.naver.com/search/all.nhn?origQuery=  xc-101       &pagingIndex=1&pagingSize=80&viewType=list&sort=rel&frm=NVSHATC&query=    xc-101a
    # https://search.shopping.naver.com/search/all.nhn?origQuery=  데니스브라운   &pagingIndex=1&pagingSize=80&viewType=list&sort=rel&frm=NVSHATC&query=    데니스브라운
    # https://search.shopping.naver.com/search/all.nhn?origQuery=  데니스브라운   &pagingIndex=1&pagingSize=80&viewType=list&sort=price_asc&frm=NVSHATC&query=   데니스브라운
    #####################################################################################################################################################################################################################################################################

class OpenMarketSpider(scrapy.Spider):
    name = "market"
    wb = None
    sheet1 = None
    fileName = None
    totalItemList = []

    # The total number of items that will be searched.
    countToSearch = 10

    # start_urls = [
    #     'https://search.shopping.naver.com/search/all.nhn?origQuery=%EC%9D%B8%EC%B2%B4%EB%AA%A8%ED%98%95%20XC-105&pagingIndex=1&pagingSize=80&viewType=list&sort=rel&frm=NVSHATC&query=%EC%9D%B8%EC%B2%B4%EB%AA%A8%ED%98%95%20XC-105',
    # ]

    def start_requests(self):
        productListEx = load_workbook(filename='가격체크리스트.xlsx')
        sheet = productListEx['Sheet1']
        self.wb = Workbook()
        self.sheet1 = self.wb.active

        # Name of xlsx file to be saved
        self.fileName = '최저가결과.xlsx'

        # Name of sheet
        self.sheet1.title = '최저가'

        for row in sheet.rows:
            currentItemID = row[0].value
            url = "https://search.shopping.naver.com/search/all.nhn?origQuery=" + currentItemID + "&pagingIndex=1&pagingSize=80&viewType=list&sort=price_asc&frm=NVSHATC&query=" + currentItemID
            yield scrapy.Request(url, self.Parse_NaverMarket)
            #             # print(productID)

        print('Total Size: ' + str(len(self.totalItemList)))
        for i in range(0, len(self.totalItemList)):
            adjustList = self.totalItemList[i][0:self.countToSearch]

            if(len(adjustList) == 0):
                continue
            else:
                for j in range(0, (len(adjustList))):
                    item = adjustList[j]

                    self.sheet1.cell(row=((i*self.countToSearch)+i)+j+1, column=1).value = item['title']
                    self.sheet1.cell(row=((i*self.countToSearch)+i)+j+1, column=2).value = item['price']
                    self.sheet1.cell(row=((i*self.countToSearch)+i)+j+1, column=3).hyperlink = item['link']

        self.wb.save(filename=self.fileName)

    # -----------------------------------------------------------------------------------------------------------------------------------------
    # Printing function.
    # This function will print the date set of item, which is called by value
    # -----------------------------------------------------------------------------------------------------------------------------------------
    def PrintInfo(self, item):
        print("|| 제품명      || " + item['title'] + "\t")
        print("|| 가격        || " + item['price'] + "\t")
        print("|| 제품링크    || " + item['link'])

        print("=======================================================================================================================================================================\n")

    # -----------------------------------------------------------------------------------------------------------------------------------------
    # Parsing function
    # -----------------------------------------------------------------------------------------------------------------------------------------
    def Parse_NaverMarket(self, response):
        items = []
        for product in response.css('li._itemSection'):
            item = ProductInfo()
            item['title'] = product.css('div.info a::attr(title)').extract_first()
            item['price'] = product.css('span.num._price_reload::text').extract_first()
            item['link'] = product.css('div.info a::attr(href)').extract_first()
            #self.PrintInfo(item)
            items.append(item)

        if(len(items) > 0):
            self.totalItemList.append(items)

        #self.QuickSort(items, 0, len(items) - 1)

    # -----------------------------------------------------------------------------------------------------------------------------------------
    # Quick Sort
    # -----------------------------------------------------------------------------------------------------------------------------------------
    def QuickSort(self, arr, low, high):
        if(low < high):
            pivot = self.Partition(arr, low, high)
            self.QuickSort(arr, low, pivot - 1)
            self.QuickSort(arr, pivot + 1, high)

    # -----------------------------------------------------------------------------------------------------------------------------------------
    # Partition function for the Quick sorting algorithm
    # -----------------------------------------------------------------------------------------------------------------------------------------
    def Partition(self, arr, low, high):
        pivotIndex = random.randint(low, high)  # Get a randomized pivot element
        pivotValue = str(arr[pivotIndex]['price'])
        pivotValue.replace(",", "")

        self.Swap(arr, pivotIndex, high)
        storeIndex = low
        for i in range(low, high):

            value = str(arr[i]['price'])
            value.replace(",", "")
            print(pivotValue)
            print(value)

            if value < pivotValue:
                self.Swap(arr, i, storeIndex)
                storeIndex += 1
        self.Swap(arr, storeIndex, high)
        return storeIndex

    #-----------------------------------------------------------------------------------------------------------------------------------------
    # Swap function for the Quick sorting algorithm
    # -----------------------------------------------------------------------------------------------------------------------------------------
    def Swap(self, arr, left, right):
        temp = arr[left]
        arr[left] = arr[right]
        arr[right] = temp

